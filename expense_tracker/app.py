import openpyxl as xl
from openpyxl.utils.exceptions import InvalidFileException
from openpyxl.chart import BarChart, reference 
from datetime import date

def process_workbook():
    while True:
        fhandle = input("Enter the name of the spreadsheet you want to open (hit 'Enter' if you want to use the default): ")
        if len(fhandle) < 1:
            workbook = xl.load_workbook('expenses.xlsx')
        else:
            try:
                workbook = xl.load_workbook(fhandle)
            except InvalidFileException:
                print("Invalid file name!")
                continue
            except FileNotFoundError:
                print("File not found!")
                continue 
        break  # when the code is executed successfully, break out of the input loop

    while True:
        sheetName = input("Enter the sheet name (or hit 'Enter' to use the default): ")
        if len(sheetName) < 1:
            sheet = workbook['Sheet1']
        else:
            try:
                sheet = workbook[sheetName]
            except KeyError:
                print("The worksheet does not exist!")
                continue
        break

    items = input("Enter the items you bought: ")
    #  validate user input for the cost of the product
    while True:
        totalCost = input("Enter the total cost: ")

        try:
            float(totalCost)
        except ValueError:
            print("The cost you entered is not a valid number!")
            continue
        break
    purchaseDate = input("Hit enter if the purchase was made today, otherwise, enter date in YYYY-MM-DD format")

    if len(purchaseDate) < 1:
        purchaseDate = date.today()
    else: 
        pass

    currMaxRow = sheet.max_row + 1  # get the max row before any insertions
    cell1 = sheet.cell(currMaxRow, 1)
    cell1.value = str(items)
    cell2 = sheet.cell(currMaxRow, 2)
    cell2.value = int(totalCost)
    cell3 = sheet.cell(currMaxRow, 3)
    cell3.value = str(purchaseDate)

    try:
        workbook.save("expenses.xlsx")
    except PermissionError:
        print("Permission denied, the file is already open!")
        

process_workbook()




